#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2022/4/21 14:21
# @Author  : WangKai
# @Site    : 
# @File    : excel.py
# @Software: PyCharm
"""
Excel
sheet_exist 装饰器 在涉及表操作之前 先检查表是否存在 没有则创建
save_excel 保存 另存为
sheet_handler 获取表对象如果表名不在excel里就创建
sheet_delete 删除表
cell_handler 获取单元格对象
query 从Excel查询指定范围的数据
row_insert 在指定的行之上插入若干行
row_delete 删除行
rows_delete_discrete 删除不连续的多行
columns_name_get 将sheet的第一行作为列名 显示所有列的名称和对应的列数
column_index_get_by_name 根据列名 返回其列号
column_get_by_col_name 指定列名 获取整一列数据
column_hidden 隐藏列
records_get 从excel查询指定范围数据 返回字典
records_write 写入符合格式数据 [{},{}]
"""
import os
import shutil
import time
from functools import wraps
from zipfile import BadZipFile
import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from commonlib.baselib.log_message import LogMessage, LOG_ERROR, LOG_DEBUG, LOG_WARN, LOG_INFO
from commonlib.baselib.msg_center import MsgCenter

MsgCenter("Excel")

def sheet_exist(func):
    """
    在涉及到表操作之前 先检查表是否存在 没有则创建
    此函数的第一个参数不能做为关键字传参
    :param func:
    :return:
    """

    @wraps(func)
    def inner(*args, **kwargs):
        LogMessage(level=LOG_INFO, module="Excel", msg=f"func {args}")
        self = args[0]
        try:
            sheet_name = args[1]
            if sheet_name not in self.sheet_list:
                LogMessage(level=LOG_INFO, module="Excel", msg=f"sheet_name:'{sheet_name}'不存在 新建....")
                self.wb.create_sheet(sheet_name)
                self.save()
                self.sheet_list.append(sheet_name)
        except IndexError:
            LogMessage(level=LOG_ERROR, module="Excel", msg=f'方法:"{func.__name__}"得一个参数请勿使用关键字传参')

        res = func(*args, **kwargs)

        return res

    return inner


class Excel:
    def __init__(self, file_name, new_flag=False):
        """
        对openpyxl封装
        :param file_name:需要操作的excel文件路径和名称
        :param new_flag:如果file_name 不存在则新建，存在也不会覆盖
        """
        self.file_name = file_name
        if os.path.exists(self.file_name):
            try:
                self.wb = openpyxl.load_workbook(file_name)
            except BadZipFile as e:
                raise BadZipFile(f"Excel 文件损坏 ,{e}.....")
        else:
            if new_flag:
                self.wb = openpyxl.Workbook()
                LogMessage(level=LOG_INFO, module="Excel", msg='sheet_name:"{}"不存在 新建~'.format(self.file_name))
            else:
                LogMessage(level=LOG_ERROR, module="Excel", msg='sheet_name:"{}"不存在'.format(self.file_name))
        # file_name 里面所有sheet的名字 (List[str])
        self.sheet_list = self.wb.sheetnames
        self.align = Alignment(horizontal="left", vertical="center")

    def save(self, backup=True):
        """
        新建或者保存excel文件
        :param backup:
        :return:
        """
        if backup:
            cur_time = time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime())
            file_name = ".".join(self.file_name.split(".")[:-1])
            suffix = self.file_name.split(".")[-1]
            print(f"{file_name}_{cur_time}.{suffix}")
            shutil.copy(self.file_name, f"{file_name}_{cur_time}.{suffix}")
        LogMessage(level=LOG_DEBUG, module="Excel", msg=f"Save '{self.file_name}'....")
        try:
            self.wb.save(self.file_name)
        except PermissionError:
            LogMessage(level=LOG_ERROR, module="Excel", msg='Data won\'t be saved !!! Close excel file ,Please')

    def sheet_delete(self, sheet_name) -> None:
        """
        删除 sheet表
        :param sheet_name:
        :return:
        """
        if sheet_name in self.sheet_list:
            sheet = self.sheet_handler(sheet_name)
            self.wb.remove(sheet)
            LogMessage(level=LOG_INFO, module="Excel", msg='Delete sheet:"{}"不存在'.format(sheet_name))
            self.save()
        else:
            LogMessage(level=LOG_ERROR, module="Excel", msg=f'Pass ,sheet: "{sheet_name}" is not in {self.sheet_list}')

    @sheet_exist
    def sheet_handler(self, sheet_name):
        """
        获取对象名 如果表不在excel里 就创建
        :param sheet_name: 表名
        :return:
        """
        return self.wb[sheet_name] if sheet_name in self.sheet_list else self.wb.create_sheet(sheet_name)

    @sheet_exist
    def cell_handler(self, sheet_name, row: int, column: int) -> Cell:
        """
        获取单元格
        :param sheet_name: 表名
        :param row: 行数
        :param column: 列数
        :return: 单元格操作对象
        """
        sheet = self.sheet_handler(sheet_name)
        sheet[f'{get_column_letter(column)}{row}'].alignment = self.align  # 设置cell格式
        return sheet.cell(row, column)

    @sheet_exist
    def query(self, sheet_name, row_start=None, row_end=None, column_start=None, column_end=None, by="row"):
        """
        从excel查询指定范围的数据
        :param sheet_name:表名
        :param row_start:开始行idx
        :param row_end:结束行idx
        :param column_start:开始列idx
        :param column_end:结束列idx
        :param by:生成迭代器
        :return:
        """
        sheet = self.sheet_handler(sheet_name)
        if by == "row":
            child_sheet = sheet.iter_rows(min_row=row_start, max_row=row_end, min_col=column_start, max_col=column_end,
                                          values_only=True)
        else:
            child_sheet = sheet.iter_cols(min_row=row_start, max_row=row_end, min_col=column_start, max_col=column_end,
                                          values_only=True)
        return child_sheet

    @sheet_exist
    def records_get(self, sheet_name, row_start=None, row_end=None, column_start=None, column_end=None,
                    by="row") -> list:
        """
        起始行/列固定为第一行 获取指定范围所有记录 返回dict-list的数据结构
        :param sheet_name:
        :param row_start:
        :param row_end:
        :param column_start:
        :param column_end:
        :param by:
        :return:
        """
        row_start = 1 if by == "row" else row_start
        column_start = 1 if by == "column" else column_start
        _iter = self.query(sheet_name, row_start=row_start, row_end=row_end, column_start=column_start,
                           column_end=column_end, by=by)
        records = [record for record in _iter]
        "将第一个record为dict 的key"
        keys = records.pop(0)
        return [dict(zip(keys, values)) for values in records]

    @sheet_exist
    def records_write(self, sheet_name, records, start_row=None) -> None:
        """
        写入符合格式数据 [{},{}...]
        :param sheet_name: 表名字
        :param records: 期望数据[{},{}...]
        :param start_row: 开始行
        :return:
        """
        sheet = self.sheet_handler(sheet_name)
        """如果传入指定行 则从指定行传入数据 否则则在最后一行插入数据"""
        row = start_row if start_row else sheet.max_row + 1
        # 判断data是否为list
        try:
            for record in records:
                # 判断元素是否为dict
                for key in record.keys():
                    self.column_index_get_by_name(sheet_name, key, write_new_col=True)
                for col_name in self.columns_name_get(sheet_name):
                    """
                    id 获取一行的key名
                    """
                    col_index = self.column_index_get_by_name(sheet_name, col_name)
                    """获取列号"""
                    value = record.get(col_name)
                    LogMessage(level=LOG_INFO, module="Excel", msg=f"In '{sheet_name}',Cell({row}, {col_name}, set{value})")
                    print(sheet_name, row, col_index)
                    self.cell_handler(sheet_name, row, col_index).value = value
                row += 1
            self.save()
            self.wb.close()
        except Exception as e:
            LogMessage(level=LOG_ERROR, module="Excel", msg=f"记录写入异常 写入未保存{e}")

    @sheet_exist
    def column_index_get_by_name(self, sheet_name, column_name, write_new_col=False) -> int:
        """
        根据列名 返回其列号
        如果没找到
        :param sheet_name:
        :param column_name:
        :param write_new_col:
        :return:
        """
        col_names = self.columns_name_get(sheet_name)
        # {'id': 1, 'dialog_id': 2, 'instance_id': 3, 'dialog_no': 4, 'action_id': 5, 'item_list': 6}
        if col_names:
            try:
                # 返回 元素的列号
                # print("sssss", col_names[column_name])
                return col_names[column_name]
            except KeyError:
                if write_new_col:
                    index = self.sheet_handler(sheet_name).max_column + 1

                    LogMessage(level=LOG_DEBUG, module="Excel",
                               msg="Column name '{}' not found ,add it at (1,{}....)".format(column_name, index))
                    self.cell_handler(sheet_name, row=1, column=index).value = column_name

                    return index
                else:
                    LogMessage(level=LOG_DEBUG, module="Excel",
                               msg="Not found '{}' in '{}'".format(column_name, col_names))
                    return 0
        # sheet 为空
        else:
            if write_new_col:
                LogMessage(level=LOG_DEBUG, module="Excel",
                           msg="Empty Sheet...,add  '{}'  at (1,1)....".format(column_name))
                self.cell_handler(sheet_name, row=1, column=1).value = column_name
                self.save()
                self.wb.close()
                return 1
            else:
                LogMessage(level=LOG_DEBUG, module="Excel",
                           msg="Not found '{}' in '{}'".format(column_name, col_names))
                return 0

    @sheet_exist
    def columns_name_get(self, sheet_name) -> dict:
        """
        将sheet第一列作为列名 显示所有的列名称对应的列数
        :param sheet_name: 表明
        :return: 所有列{名称：列标}
        """
        sheet = self.sheet_handler(sheet_name)
        try:
            columns_names = [cell.value for cell in list(sheet.rows)[0]]
            # {'id': 1, 'dialog_id': 2, 'instance_id': 3, 'dialog_no': 4, 'action_id': 5, 'item_list': 6}
            return {k: v + 1 for v, k in enumerate(columns_names)}
        except IndexError:
            LogMessage(level=LOG_ERROR, module="Excel", msg="Empty Sheet.....")
            return dict()
